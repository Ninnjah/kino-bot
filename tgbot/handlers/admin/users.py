import logging
from datetime import datetime
from io import BytesIO

from aiogram import Router, F
from aiogram.types import CallbackQuery, BufferedInputFile

from aiogram_dialog import DialogManager, Dialog, Window
from aiogram_dialog.widgets.text import Format, List
from aiogram_dialog.widgets.kbd import (
    Row,
    FirstPage,
    PrevPage,
    CurrentPage,
    NextPage,
    LastPage,
    Button,
)

from openpyxl import Workbook

from tgbot.handlers.admin.states.users import UserSG
from tgbot.services.l10n_dialog import L10NFormat
from tgbot.services.repository import Repo

logger = logging.getLogger(__name__)
router = Router(name=__name__)

PAGE_SIZE = 20


async def get_users(dialog_manager: DialogManager, repo: Repo, **kwargs):
    user_list = await repo.list_users()

    users = []
    for num, user in enumerate(user_list, start=1):
        users.append(
            "{num}. <a href='tg://user?id={user_id}'><b>{user_id}</b></a> [{date}]".format(
                num=num,
                user_id=user.id,
                date=user.created_on.strftime("%Y.%m.%d %H:%M"),
            )
        )

    return {"users": users}


async def export_users(callback: CallbackQuery, button: Button, manager: DialogManager):
    repo: Repo = manager.middleware_data["repo"]
    users = await repo.list_users()

    wb = Workbook()
    ws = wb.active

    ws.row_dimensions[1].height = 30
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 20

    ws["A1"] = "ID"
    ws["B1"] = "Имя"
    ws["C1"] = "Фамилия"
    ws["D1"] = "Никнейм"
    ws["E1"] = "Запись создана"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for user in users:
        ws.append(
            [
                user.id,
                user.firstname,
                user.lastname,
                user.username,
                user.created_on,
            ]
        )

    # Save the file
    with BytesIO() as table:
        wb.save(table)
        table.seek(0)
        await callback.message.answer_document(
            BufferedInputFile(
                table.read(),
                filename=f"users.{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            )
        )


user_list_dialog = Dialog(
    Window(
        L10NFormat("admin-users-list"),
        List(
            Format("{item}"),
            items="users",
            id="users_scroll",
            page_size=PAGE_SIZE,
        ),
        Row(
            FirstPage(scroll="users_scroll"),
            PrevPage(scroll="users_scroll"),
            CurrentPage(scroll="users_scroll"),
            NextPage(scroll="users_scroll"),
            LastPage(scroll="users_scroll"),
            when=F["users"].len() > PAGE_SIZE,
        ),
        Button(
            L10NFormat("admin-users-export"),
            id="export",
            on_click=export_users,
        ),
        getter=get_users,
        state=UserSG.lst,
    ),
)


router.include_routers(
    user_list_dialog,
)
